import os
import sqlite3
import logging
import secrets
from flask import Flask, redirect, request, render_template, send_file, jsonify, session, url_for
from io import BytesIO, StringIO
from datetime import datetime, timedelta
import pandas as pd  # Для экспорта в Excel
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

# -----------------------------
# Определяем базовую директорию и путь к БД
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, 'phishing_data.db')

# -----------------------------
# Инициализация Flask-приложения
# -----------------------------
app = Flask(__name__, template_folder=BASE_DIR)

# -----------------------------
# НАСТРОЙКИ СЕССИЙ И БЕЗОПАСНОСТИ
# -----------------------------
app.secret_key = 'phishing-dashboard-2026-super-secret-key'  # Ключ для шифрования сессий

# Таймаут неактивной сессии (автовыход)
IDLE_TIMEOUT_MINUTES = 15
app.permanent_session_lifetime = timedelta(minutes=IDLE_TIMEOUT_MINUTES)

# Настройки cookies
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=False  # True при HTTPS
)

# -----------------------------
# Настройка логирования
# -----------------------------
class IgnoreApiClicksFilter(logging.Filter):
    """Фильтрует GET запросы к /api/clicks и /api/emails из логов werkzeug"""
    def filter(self, record):
        message = str(record.getMessage())
        # Фильтруем все варианты GET запросов к API эндпоинтам
        api_patterns = ['/api/clicks', '/api/emails']
        for pattern in api_patterns:
            # Проверяем различные форматы логов werkzeug
            if (f'GET {pattern}' in message or
                f'{pattern} HTTP' in message or
                f'{pattern}" 200' in message or
                f'{pattern}" 304' in message):
                return False  # Не логируем
        return True  # Логируем остальное

werkzeug_logger = logging.getLogger('werkzeug')
werkzeug_logger.addFilter(IgnoreApiClicksFilter())
werkzeug_logger.setLevel(logging.INFO)

# -----------------------------
# DB init
# -----------------------------
def initialize_db():
    """Функция для инициализации базы данных, если она еще не существует"""
    if not os.path.exists(DATABASE):
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE clicks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                ip TEXT,
                user_agent TEXT,
                token TEXT
            )
        ''')
        conn.commit()
        conn.close()

# -----------------------------
# Таблица пользователей
# -----------------------------
def initialize_users():
    """Создание таблицы users и дефолтного администратора"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TEXT
        )
    ''')

    # Если пользователей нет — создаем admin:passwd123
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
            (
                'admin',
                generate_password_hash('passwd123'),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            )
        )

    conn.commit()
    conn.close()

# -----------------------------
# Таблица email адресов
# -----------------------------
def initialize_emails():
    """Создание таблицы emails для хранения email адресов"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS emails (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            token TEXT UNIQUE NOT NULL,
            created_at TEXT
        )
    ''')
    
    # Проверяем, существует ли колонка token
    cursor.execute("PRAGMA table_info(emails)")
    columns = [col[1] for col in cursor.fetchall()]
    
    if 'token' not in columns:
        # Добавляем колонку token без UNIQUE сначала
        cursor.execute("ALTER TABLE emails ADD COLUMN token TEXT")
        conn.commit()
    
    # Генерируем токены для записей без токенов
    cursor.execute("SELECT id, email FROM emails WHERE token IS NULL OR token = ''")
    rows_without_token = cursor.fetchall()
    
    import hashlib
    import secrets
    
    for row_id, email in rows_without_token:
        # Генерируем уникальный токен на основе email и id
        unique_token = None
        max_attempts = 10
        attempt = 0
        
        while unique_token is None and attempt < max_attempts:
            # Генерируем токен на основе email и случайного значения
            token_base = f"{email}_{row_id}_{secrets.token_hex(8)}_{datetime.now().timestamp()}"
            token_hash = hashlib.md5(token_base.encode()).hexdigest()[:16]
            
            # Проверяем уникальность
            cursor.execute("SELECT COUNT(*) FROM emails WHERE token = ?", (token_hash,))
            if cursor.fetchone()[0] == 0:
                unique_token = token_hash
            attempt += 1
        
        if unique_token:
            cursor.execute("UPDATE emails SET token = ? WHERE id = ?", (unique_token, row_id))
    
    # Если колонка была добавлена, добавляем UNIQUE constraint через пересоздание таблицы
    if 'token' not in columns and rows_without_token:
        conn.commit()
        # Создаем новую таблицу с UNIQUE constraint
        cursor.execute('''
            CREATE TABLE emails_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                token TEXT UNIQUE NOT NULL,
                created_at TEXT
            )
        ''')
        cursor.execute("INSERT INTO emails_new SELECT id, email, token, created_at FROM emails")
        cursor.execute("DROP TABLE emails")
        cursor.execute("ALTER TABLE emails_new RENAME TO emails")

    conn.commit()
    conn.close()

# -----------------------------
# CSRF
# -----------------------------
def generate_csrf_token():
    """Генерируем CSRF токен и сохраняем в сессии"""
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

app.jinja_env.globals['csrf_token'] = generate_csrf_token

@app.before_request
def csrf_protect():
    """Проверка CSRF токена при POST запросах"""
    if request.method == "POST":
        # Пропускаем проверку для загрузки файлов (multipart/form-data)
        # если это не форма с CSRF токеном
        if request.path == '/api/emails/upload' and 'file' in request.files:
            # Для загрузки файлов проверяем только заголовок
            token = session.get('_csrf_token')
            form_token = request.headers.get('X-CSRF-Token')
            if not token or token != form_token:
                return "CSRF validation failed", 403
        else:
            # Для обычных POST запросов
            token = session.get('_csrf_token')
            form_token = request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token')
            if not token or token != form_token:
                return "CSRF validation failed", 403

# -----------------------------
# Автовыход по неактивности
# -----------------------------
@app.before_request
def session_timeout_check():
    """Если пользователь авторизован, обновляем сессию, чтобы отслеживать активность"""
    if session.get('user_id'):
        session.permanent = True  # Используем timedelta из app.permanent_session_lifetime
        session.modified = True

# -----------------------------
# Декоратор авторизации
# -----------------------------
def login_required(f):
    """Обертка для маршрутов, требующих авторизации"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# -----------------------------
# Авторизация
# -----------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Маршрут для логина"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, password_hash FROM users WHERE username = ?",
            (username,)
        )
        user = cursor.fetchone()
        conn.close()

        if user and check_password_hash(user[1], password):
            session.clear()
            session['user_id'] = user[0]
            session['username'] = username
            return redirect(url_for('dashboard'))

        return render_template('login.html', error='Неверный логин или пароль')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Выход из аккаунта"""
    session.clear()
    return redirect(url_for('login'))

# -----------------------------
# Чтение данных из базы
# -----------------------------
def get_report_data():
    """Получаем все клики из БД"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM clicks ORDER BY id DESC")
    data = cursor.fetchall()
    conn.close()

    return [
        {
            "id": row[0],
            "timestamp": row[1],
            "ip": row[2],
            "user_agent": row[3],
            "token": row[4]
        }
        for row in data
    ]

# -----------------------------
# Экспорт данных
# -----------------------------
@app.route('/export/excel')
@login_required
def export_excel():
    """Экспорт в Excel"""
    df = pd.DataFrame(get_report_data())
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='phishing_report.xlsx')

@app.route('/export/txt')
@login_required
def export_txt():
    """Экспорт в TXT"""
    buffer = StringIO()
    for row in get_report_data():
        buffer.write(f"{row['timestamp']} | {row['ip']} | {row['user_agent']} | {row['token']}\n")
    return send_file(BytesIO(buffer.getvalue().encode()), as_attachment=True, download_name='phishing_report.txt')

@app.route('/export/sql')
@login_required
def export_sql():
    """Экспорт в SQL"""
    buffer = StringIO()
    buffer.write("CREATE TABLE clicks (timestamp TEXT, ip TEXT, user_agent TEXT, token TEXT);\n\n")
    for row in get_report_data():
        buffer.write(
            f"INSERT INTO clicks VALUES ('{row['timestamp']}', '{row['ip']}', '{row['user_agent']}', '{row['token']}');\n"
        )
    return send_file(BytesIO(buffer.getvalue().encode()), as_attachment=True, download_name='phishing_report.sql')

@app.route('/export/db')
@login_required
def export_db():
    """Экспорт базы SQLite"""
    buffer = BytesIO()
    conn = sqlite3.connect(':memory:')
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE clicks (timestamp TEXT, ip TEXT, user_agent TEXT, token TEXT)")
    for row in get_report_data():
        cursor.execute("INSERT INTO clicks VALUES (?, ?, ?, ?)",
                       (row['timestamp'], row['ip'], row['user_agent'], row['token']))
    for line in conn.iterdump():
        buffer.write(f"{line}\n".encode())
    conn.close()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='phishing_report.db')

# -----------------------------
# API
# -----------------------------
@app.route('/api/clicks')
@login_required
def api_clicks():
    """Возвращает все клики в JSON"""
    return jsonify(get_report_data())

@app.route('/api/delete', methods=['POST'])
@login_required
def api_delete():
    """Удаление выбранных записей"""
    ids = request.json.get('ids', [])
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute(f"DELETE FROM clicks WHERE id IN ({','.join('?' for _ in ids)})", ids)
    conn.commit()
    conn.close()
    return {"status": "ok"}

# -----------------------------
# Email Management
# -----------------------------
@app.route('/emails')
@login_required
def emails():
    """Страница управления email адресами"""
    return render_template('emails.html')

@app.route('/api/emails', methods=['GET', 'POST'])
@login_required
def api_emails():
    """Получение списка email или добавление новых"""
    if request.method == 'GET':
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT id, email, token, created_at FROM emails ORDER BY id DESC")
        data = cursor.fetchall()
        conn.close()
        
        # Получаем базовый URL для ссылок
        base_url = request.host_url.rstrip('/')
        if not base_url.startswith('http'):
            base_url = f"http://{request.host}"

        return jsonify([
            {
                "id": row[0],
                "email": row[1],
                "token": row[2] or '',
                "tracking_link": f"{base_url}/track/{row[2]}" if row[2] else '',
                "created_at": row[3]
            }
            for row in data
        ])
    
    elif request.method == 'POST':
        # Добавление email адресов
        emails_text = request.json.get('emails', '')
        if not emails_text:
            return jsonify({"error": "Список email пуст"}), 400
        
        # Разбиваем на строки и по запятым, затем очищаем
        email_list = []
        # Сначала разбиваем по переносам строк
        for line in emails_text.split('\n'):
            # Затем разбиваем каждую строку по запятым
            for email in line.split(','):
                email = email.strip()
                if email:
                    email_list.append(email)
        
        if not email_list:
            return jsonify({"error": "Нет валидных email адресов"}), 400
        
        # Простая валидация email
        import re
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        
        added = 0
        skipped = 0
        errors = []
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Импортируем для генерации токенов
        import hashlib
        import secrets
        
        for email in email_list:
            # Проверка формата
            if not email_pattern.match(email):
                errors.append(f"Неверный формат: {email}")
                skipped += 1
                continue
            
            # Генерируем уникальный токен
            unique_token = None
            max_attempts = 10
            attempt = 0
            
            while unique_token is None and attempt < max_attempts:
                # Генерируем токен на основе email и случайного значения
                token_base = f"{email.lower()}_{secrets.token_hex(8)}_{datetime.now().timestamp()}"
                token_hash = hashlib.md5(token_base.encode()).hexdigest()[:16]
                
                # Проверяем уникальность
                cursor.execute("SELECT COUNT(*) FROM emails WHERE token = ?", (token_hash,))
                if cursor.fetchone()[0] == 0:
                    unique_token = token_hash
                attempt += 1
            
            if unique_token is None:
                errors.append(f"Не удалось сгенерировать уникальный токен для: {email}")
                skipped += 1
                continue
            
            # Проверка уникальности и добавление
            try:
                cursor.execute(
                    "INSERT INTO emails (email, token, created_at) VALUES (?, ?, ?)",
                    (email.lower(), unique_token, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                )
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
                errors.append(f"Дубликат: {email}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "status": "ok",
            "added": added,
            "skipped": skipped,
            "errors": errors
        })

@app.route('/api/emails/delete', methods=['POST'])
@login_required
def api_emails_delete():
    """Удаление выбранных email адресов"""
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({"error": "Не выбраны email для удаления"}), 400
    
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute(f"DELETE FROM emails WHERE id IN ({','.join('?' for _ in ids)})", ids)
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

@app.route('/api/emails/upload', methods=['POST'])
@login_required
def api_emails_upload():
    """Загрузка email адресов из файла (xlsx, txt, csv)"""
    if 'file' not in request.files:
        return jsonify({"error": "Файл не загружен"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "Файл не выбран"}), 400
    
    try:
        email_list = []
        
        # Обработка xlsx файлов
        if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
            # Используем openpyxl напрямую для чтения сырых значений из Excel
            # Это позволяет получить оригинальные значения до любых преобразований
            try:
                from openpyxl import load_workbook
                import io
                
                # Сохраняем позицию файла и читаем его содержимое
                file.seek(0)
                file_content = file.read()
                file.seek(0)
                
                # Загружаем workbook через openpyxl
                wb = load_workbook(io.BytesIO(file_content), data_only=False)
                ws = wb.active
                
                # Ищем колонку с email в первой строке
                email_col_idx = None
                header_row = 1
                has_header = False
                
                # Проверяем первую строку - есть ли там заголовок
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    cell_value = str(cell.value).lower() if cell.value else ''
                    if 'email' in cell_value or 'mail' in cell_value or 'e-mail' in cell_value:
                        email_col_idx = col_idx
                        has_header = True
                        break
                
                # Если не нашли колонку с email, берем первую колонку
                if email_col_idx is None:
                    email_col_idx = 1
                    # Проверяем, является ли первая строка заголовком или данными
                    first_cell = ws.cell(row=1, column=email_col_idx).value
                    if first_cell and '@' in str(first_cell):
                        # Первая строка содержит email, значит заголовка нет
                        has_header = False
                    else:
                        has_header = True
                
                # Определяем с какой строки начинать чтение данных
                start_row = 2 if has_header else 1
                
                # Читаем все значения из колонки
                for row_num in range(start_row, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=email_col_idx)
                    if cell.value is None:
                        continue
                    
                    # Получаем значение как строку
                    # Если это число (Excel интерпретировал как число), пытаемся восстановить
                    email_str = None
                    
                    if cell.data_type == 'n':  # число
                        # Если это число, но должно быть email, возможно Excel неправильно интерпретировал
                        # Проверяем, есть ли @ в отображаемом значении
                        display_value = str(cell.value)
                        if '@' in display_value:
                            email_str = display_value
                        else:
                            # Это чистое число без @, пропускаем
                            continue
                    else:
                        # Это текст или другой тип
                        email_str = str(cell.value).strip()
                    
                    # Дополнительная проверка на пустые значения
                    if not email_str or email_str.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Если это число без @ символа, пропускаем
                    if '@' not in email_str:
                        try:
                            float(email_str)
                            continue  # Это число, не email
                        except (ValueError, TypeError):
                            pass  # Это не число, продолжаем
                    
                    email_list.append(email_str)
                
                wb.close()
                
            except (ImportError, Exception) as e:
                # Если openpyxl недоступен или произошла ошибка, используем pandas
                file.seek(0)
                
                # Читаем Excel с принудительным преобразованием всех значений в строки
                df = pd.read_excel(file, dtype=str, keep_default_na=False, na_filter=False)
                
                # Ищем колонку с email
                email_column = None
                for col in df.columns:
                    col_lower = str(col).lower()
                    if 'email' in col_lower or 'mail' in col_lower or 'e-mail' in col_lower:
                        email_column = col
                        break
                
                if email_column is None:
                    email_column = df.columns[0]
                
                # Извлекаем email адреса
                for value in df[email_column]:
                    email_str = str(value).strip()
                    
                    # Пропускаем пустые значения и числа без @
                    if not email_str or email_str.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Если это число без @ символа, пропускаем
                    if '@' not in email_str:
                        try:
                            float(email_str)
                            continue  # Это число, не email
                        except (ValueError, TypeError):
                            pass  # Это не число, продолжаем
                    
                    email_list.append(email_str)
        
        # Обработка текстовых файлов
        elif file.filename.endswith('.txt') or file.filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            for line in content.split('\n'):
                # Разбиваем по запятым и переносам строк
                for email in line.split(','):
                    email = email.strip()
                    if email:
                        email_list.append(email)
        else:
            return jsonify({"error": "Неподдерживаемый формат файла. Используйте .xlsx, .txt или .csv"}), 400
        
        if not email_list:
            return jsonify({"error": "В файле не найдено email адресов"}), 400
        
        # Валидация и добавление в БД
        import re
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        
        added = 0
        skipped = 0
        errors = []
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Импортируем для генерации токенов
        import hashlib
        import secrets
        
        for email in email_list:
            # Проверка формата
            if not email_pattern.match(email):
                errors.append(f"Неверный формат: {email}")
                skipped += 1
                continue
            
            # Генерируем уникальный токен
            unique_token = None
            max_attempts = 10
            attempt = 0
            
            while unique_token is None and attempt < max_attempts:
                # Генерируем токен на основе email и случайного значения
                token_base = f"{email.lower()}_{secrets.token_hex(8)}_{datetime.now().timestamp()}"
                token_hash = hashlib.md5(token_base.encode()).hexdigest()[:16]
                
                # Проверяем уникальность
                cursor.execute("SELECT COUNT(*) FROM emails WHERE token = ?", (token_hash,))
                if cursor.fetchone()[0] == 0:
                    unique_token = token_hash
                attempt += 1
            
            if unique_token is None:
                errors.append(f"Не удалось сгенерировать уникальный токен для: {email}")
                skipped += 1
                continue
            
            # Проверка уникальности и добавление
            try:
                cursor.execute(
                    "INSERT INTO emails (email, token, created_at) VALUES (?, ?, ?)",
                    (email.lower(), unique_token, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                )
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
                errors.append(f"Дубликат: {email}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "status": "ok",
            "added": added,
            "skipped": skipped,
            "errors": errors
        })
    
    except Exception as e:
        return jsonify({"error": f"Ошибка при обработке файла: {str(e)}"}), 500

# -----------------------------
# Панель управления
# -----------------------------
@app.route('/')
@login_required
def dashboard():
    """Главная страница dashboard"""
    return render_template('dashboard.html')

# -----------------------------
# Трекинг (ПУБЛИЧНЫЙ)
# -----------------------------
@app.route('/track/<token>')
def track(token):
    """Публичный маршрут для трекинга кликов по токену"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO clicks (timestamp, ip, user_agent, token) VALUES (?, ?, ?, ?)",
        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
         request.remote_addr,
         request.headers.get('User-Agent', 'unknown'),
         token)
    )
    conn.commit()
    conn.close()
    return redirect('https://ya.ru')

# -----------------------------
# Запуск
# -----------------------------
if __name__ == '__main__':
    initialize_db()
    initialize_users()
    initialize_emails()
    app.run(host='0.0.0.0', port=8080, debug=False, use_reloader=False)
